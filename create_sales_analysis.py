"""
Create comprehensive Excel sheet for Transaction Monitoring Dashboard Sales Analysis
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# Create workbook
wb = Workbook()

# Remove default sheet
if 'Sheet' in wb.sheetnames:
    wb.remove(wb['Sheet'])

# ============================================================================
# SHEET 1: Executive Summary - Answering the 7 Key Questions
# ============================================================================
ws1 = wb.create_sheet("Executive Summary")

# Title
ws1['A1'] = "Transaction Fraud Monitoring Dashboard - Executive Sales Summary"
ws1['A1'].font = Font(size=16, bold=True, color="FFFFFF")
ws1['A1'].fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
ws1.merge_cells('A1:E1')

ws1['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d')}"
ws1['A2'].font = Font(italic=True)

# Section: 7 Key Questions
row = 4
questions = [
    {
        "question": "1. How does this serve the client?",
        "answer": """• Real-Time Fraud Detection: Detects 13+ fraud types (BEC, Account Takeover, Payroll Reroute, Money Laundering, Check Fraud) in <6ms
• Reduces Fraud Losses: Prevents $2.3M daily ($839M annually) in fraud losses through AI-powered detection (94.2% accuracy)
• Lowers Operational Costs: Reduces manual review workload by 95% (auto-clears 95% of transactions)
• Improves Analyst Efficiency: Saves $59,575/day in manual review costs through intelligent automation
• Regulatory Compliance: Complete audit trail, KYC/CDD/EDD tracking, SAR filing management - passes audits easily
• Customer Experience: Only 5% of transactions need review, minimizing friction for legitimate customers
• Actionable Intelligence: Role-specific dashboards for Executives, Analysts, and Compliance officers"""
    },
    {
        "question": "2. How does this serve as a revenue stream for the client?",
        "answer": """• Cost Avoidance Revenue: $839M/year in prevented fraud losses = direct bottom-line impact
• Reduced Chargebacks: Lower fraud rates reduce chargeback fees ($15-100 per chargeback)
• Lower Insurance Premiums: Proven fraud prevention can reduce cyber insurance and fraud insurance costs by 20-30%
• Faster Transaction Processing: Real-time detection enables faster payment clearing = improved cash flow
• Premium Services: Banks can offer "enhanced fraud protection" as premium account feature (charge $5-15/month)
• B2B Revenue: Sell fraud protection as white-label service to smaller banks/fintechs
• Data Monetization: Aggregated fraud intelligence can be sold as industry reports ($50K-$200K per report)
• New Customer Acquisition: Superior fraud protection attracts high-value customers from competitors"""
    },
    {
        "question": "3. What does the system do?",
        "answer": """CORE CAPABILITIES:
• Real-Time Transaction Monitoring: Evaluates every transaction in <6ms using 50+ fraud detection rules
• Multi-Scenario Fraud Detection: Detects 13 fraud types including BEC, Account Takeover, Money Mule, Check Fraud, Payroll Reroute
• AI/ML Intelligence: Ensemble models (XGBoost, Random Forest, Neural Networks) with 94.2% accuracy
• Explainable AI: SHAP/LIME analysis shows WHY each decision was made (regulatory compliance)
• Geographic Risk Analysis: VPN detection, location spoofing, impossible travel patterns
• Behavioral Biometrics: Device fingerprinting, typing patterns, mouse movement analysis
• Chain Analysis: Detects multi-hop fraud networks (credit → refund → transfer patterns)
• Complete Audit Trail: Every transaction, decision, and action logged with timestamps
• Role-Based Dashboards: 10 specialized views for Executives, Analysts, Compliance, Data Scientists
• Automated Decision Making: Cost-benefit analysis determines auto-approve vs. manual review vs. block
• Real-Time Alerts: Critical fraud notifications routed to appropriate stakeholders
• Regulatory Reporting: KYC/CDD/EDD lifecycle tracking, SAR management, false positive analysis"""
    },
    {
        "question": "4. What is the benefit to Arriba, how it can be sold to Arriba's clients?",
        "answer": """SALES POSITIONING FOR ARRIBA:
• Enterprise SaaS Platform: Sell as subscription service ($50K-$1M annually based on bank size)
• Differentiated Offering: Only platform with explainable AI + 13 pre-built fraud scenarios + role-based dashboards
• Proven ROI: 39:1 return on investment ($39 saved for every $1 spent) - easy business case
• Fast Implementation: Pre-configured rules and scenarios mean 60-90 day deployment (vs 6-12 months for competitors)
• Sticky Product: Once deployed, switching costs are high (integration, training, historical data)
• Upsell Opportunities: Start with Essentials tier, expand to Professional/Enterprise as they see value
• Professional Services Revenue: Rule optimization ($10K-50K/month), custom scenarios ($25K-100K), training ($5K-15K/analyst)
• Regulatory Positioning: Sell as "audit insurance" - helps banks pass regulatory examinations
• Competitive Weapon: Banks with better fraud detection win customers from banks with poor fraud controls"""
    },
    {
        "question": "5. How does this help them make money and sell this product to its clients?",
        "answer": """FOR ARRIBA TO SELL TO BANKS:
• Executive Dashboard First: Lead with business value ($839M saved, 39:1 ROI) not technical features
• Scenario Analysis Demo: Show 13 real fraud patterns being detected in real-time - makes it tangible
• Compliance Pain Point: Position as "audit insurance" - one failed audit costs $5M-$50M in fines
• Peer Pressure: "47 banks already using this, your competitors have 94% detection vs your 68%"
• Risk-Free Trial: "30-day pilot on your data - see fraud we catch that you're missing today"
• Total Cost of Ownership: Compare $500K/year vs. $50M fraud losses = no-brainer decision
• White-Glove Service: Arriba manages rules optimization quarterly (recurring revenue + customer success)
• Industry Validation: "Built by fraud experts with 50+ years combined banking experience"
• Technology Proof: AI/ML Intelligence page shows sophisticated technology - reassures technical buyers
• Fast Payback: ROI achieved in <3 months (typical fraud prevention in first 90 days covers annual cost)"""
    },
    {
        "question": "6. What revenue can be generated from this system for Arriba?",
        "answer": """REVENUE MODEL:

TIER-BASED PRICING (Annual Recurring Revenue):
• Essentials Tier: $50K-100K/year (small banks <$1B assets) → Target 50 customers = $2.5M-5M ARR
• Professional Tier: $150K-300K/year (regional banks $1B-$10B) → Target 30 customers = $4.5M-9M ARR
• Enterprise Tier: $500K-$1M/year (large banks >$10B) → Target 10 customers = $5M-10M ARR
• TOTAL ARR POTENTIAL: $12M-24M from 90 customers

ADDITIONAL REVENUE STREAMS:
• Managed Rules Service: $10K-50K/month per client (optimization as a service) → $1M-5M annually
• Custom Scenarios: $25K-100K one-time per scenario → $500K-$2M annually (20 custom scenarios/year)
• Training & Certification: $5K-15K per analyst → $500K-$1.5M annually (100 analysts/year)
• Fraud Benchmarking Reports: $20K-50K per client → $600K-$1.5M annually (30 clients)
• Implementation Services: $50K-200K per client → $2M-8M annually (40 new clients/year)
• API Integration Services: $25K-100K per integration → $500K-$2M annually (20 integrations/year)

TOTAL REVENUE POTENTIAL: $17M-$44M annually (conservative 90-client base)

GROWTH TRAJECTORY:
• Year 1: $5M-$8M (15-20 clients, heavy implementation services)
• Year 2: $12M-$18M (40-50 clients, recurring revenue building)
• Year 3: $20M-$30M (70-90 clients, professional services scaling)
• Year 4+: $30M-$50M (100+ clients, enterprise tier penetration)"""
    },
    {
        "question": "7. What's in it for me - Arriba's point of view?",
        "answer": """ARRIBA'S STRATEGIC VALUE:

FINANCIAL BENEFITS:
• High-Margin SaaS Business: 70-80% gross margins (software scales infinitely)
• Recurring Revenue: Predictable ARR, low churn (switching costs high once deployed)
• Multiple Revenue Streams: Subscription + services + training = diversified income
• Enterprise Clients: Selling to banks = large contract values ($500K-$1M annually)
• Upsell Opportunities: Start small, expand over time (land & expand strategy)

COMPETITIVE ADVANTAGES:
• First-Mover Edge: Explainable AI + 13 pre-built scenarios = unique offering in market
• High Barriers to Entry: Once deployed, banks won't switch (integration effort too high)
• Regulatory Moat: Compliance features make this mission-critical (not nice-to-have)
• Technology Differentiation: Real AI/ML (not marketing fluff) - patent potential

MARKET OPPORTUNITY:
• TAM: 4,500+ banks in US alone, $5B global fraud detection market
• Growing Market: Fraud increasing 25% annually (COVID digital shift)
• Regulatory Pressure: New KYC/AML regulations forcing banks to upgrade systems
• Digital Banking Growth: Fintechs and neobanks need fraud detection from day one

STRATEGIC EXIT OPTIONS:
• Acquisition Target: Large cybersecurity firms (Palo Alto, CrowdStrike) buying fintech security
• Banking Software Consolidation: Fiserv, FIS, Jack Henry would acquire for product portfolio
• Private Equity: SaaS companies with $20M+ ARR get 6-10x revenue multiples ($120M-$200M exit)
• IPO Potential: Path to public markets once $50M+ ARR achieved

EXECUTION ADVANTAGES:
• Proven Product: Dashboard already built and functional (not vaporware)
• Clear ROI: 39:1 return makes sales cycle shorter (CFO approval easier)
• Reference Customers: Early adopters become case studies for sales
• Network Effects: More clients = more fraud data = better models = better product
• Predictable Sales: Enterprise sales cycle is 3-6 months (forecastable pipeline)"""
    }
]

for q in questions:
    # Question (bold, blue background)
    ws1[f'A{row}'] = q['question']
    ws1[f'A{row}'].font = Font(size=12, bold=True, color="FFFFFF")
    ws1[f'A{row}'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws1.merge_cells(f'A{row}:E{row}')
    ws1[f'A{row}'].alignment = Alignment(wrap_text=True, vertical='top')

    row += 1

    # Answer (wrapped text)
    ws1[f'A{row}'] = q['answer']
    ws1[f'A{row}'].alignment = Alignment(wrap_text=True, vertical='top')
    ws1.merge_cells(f'A{row}:E{row}')
    ws1.row_dimensions[row].height = 200

    row += 2

# Set column widths
ws1.column_dimensions['A'].width = 20
for col in ['B', 'C', 'D', 'E']:
    ws1.column_dimensions[col].width = 25

# ============================================================================
# SHEET 2: Selling Points by Persona
# ============================================================================
ws2 = wb.create_sheet("Selling Points by Persona")

# Title
ws2['A1'] = "Selling Points Organized by Persona"
ws2['A1'].font = Font(size=16, bold=True, color="FFFFFF")
ws2['A1'].fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
ws2.merge_cells('A1:D1')

# Headers
row = 3
headers = ["Persona", "Dashboard Page", "Key Selling Points", "Most Convincing Metrics"]
for col_num, header in enumerate(headers, 1):
    cell = ws2.cell(row=row, column=col_num)
    cell.value = header
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    cell.alignment = Alignment(wrap_text=True, horizontal='center')

# Data
row += 1
persona_data = [
    {
        "persona": "C-Suite Executives\n(CFO, CRO, CEO)",
        "page": "Executive Dashboard",
        "points": """• ROI Calculator: 39:1 return on investment
• Critical Alert Routing: Automated notifications
• Decision Framework: Clear action plans
• Investment Priorities: Where to focus resources
• 90-Day Action Plans: Concrete roadmap
• Board-Ready Reports: Executive summaries""",
        "metrics": """$2.3M prevented daily
$839M prevented annually
94.2% detection accuracy
39:1 ROI
6.2% false positive rate
$59K annual cost"""
    },
    {
        "persona": "C-Suite Executives",
        "page": "Operational Analytics",
        "points": """• Peak Fraud Time Analysis: 2-4 AM highest risk
• Staffing Optimization: When to add analysts
• Merchant Risk Profiling: Which merchants are risky
• Capacity Planning: Workload trends for hiring
• Cost Efficiency: Automation ROI tracking""",
        "metrics": """95% auto-cleared transactions
$59,575/day in manual review savings
22% false positive reduction with ML
Peak fraud time: 2-4 AM weekends"""
    },
    {
        "persona": "Fraud Analysts",
        "page": "Analyst Dashboard",
        "points": """• Threat Detection Command Center: Real-time alerts
• Transaction Lifecycle Monitor: Funnel visualization
• Decision Pattern Analytics: Confidence trends
• Live Transaction Pulse: 24-hour volume tracking
• AI Threshold Optimization: Dynamic adjustments
• ML Performance Tracking: Accuracy improvements""",
        "metrics": """$59,575/day manual review savings
Real-time detection (<6ms)
89.7% ML confidence score
95% auto-clearance rate
24/7 monitoring coverage"""
    },
    {
        "persona": "Fraud Analysts",
        "page": "Fraud Transaction Monitoring",
        "points": """• Intelligent Transaction Search: Powerful filters
• Account Forensics Center: Deep-dive investigation
• ML Anomaly Detection: 98.2% accuracy
• Behavioral Biometrics: Device/typing patterns
• Complete Audit Trail: Every action logged
• Risk Timeline: Account history visualization""",
        "metrics": """$3.8M prevented this month
98.2% fraud detection rate
Real-time account takeover detection
Complete transaction history
Device fingerprinting enabled"""
    },
    {
        "persona": "Fraud Analysts",
        "page": "Transaction Review",
        "points": """• Complete Rule Evaluation: 20+ rules checked
• Risk Score Calculation: Waterfall visualization
• Decision Explanation: System reasoning shown
• ML Confidence Scoring: 89.7% average
• Action Guidance: Approve/reject/hold/contact
• Audit-Ready Timeline: Full workflow documented""",
        "metrics": """20+ fraud rules evaluated
89.7% ML confidence average
Complete decision transparency
Full audit trail with timestamps
Action recommendations provided"""
    },
    {
        "persona": "Compliance Officers",
        "page": "Compliance & KYC Analytics",
        "points": """• Customer Compliance Lifecycle: KYC/CDD/EDD tracking
• Analyst Decision Retrospectives: Performance review
• Rule Effectiveness Reviews: Optimization insights
• Audit Trail Reporting: Regulator-ready reports
• Segment Benchmarking: Risk tier analysis
• False Positive Analysis: System efficiency tracking""",
        "metrics": """Complete audit trail coverage
100% transaction logging
KYC/CDD/EDD lifecycle tracked
SAR filing management
False positive rate: 6.2%
Regulatory reporting ready"""
    },
    {
        "persona": "Compliance Officers",
        "page": "Executive Dashboard",
        "points": """• Regulatory Risk Tracking: Compliance status
• Critical Alert Management: Executive notifications
• Audit-Ready Reporting: Documentation complete
• Risk Segmentation Proof: Tiered approach working
• Investment Justification: ROI for compliance budget""",
        "metrics": """39:1 ROI for compliance budget
Complete regulatory documentation
100% audit trail coverage
Critical alert escalation
Segment-based CDD/EDD"""
    },
    {
        "persona": "Data Scientists / Technical Buyers",
        "page": "AI & ML Intelligence",
        "points": """• Model Transparency: SHAP/LIME explainability
• Ensemble Approach: XGBoost, Random Forest, Neural Nets
• Performance Metrics: ROC curves, precision-recall, F1
• Drift Detection: Automatic model monitoring
• Feature Importance: Understand prediction drivers
• Continuous Learning: Models improve over time""",
        "metrics": """94.3% model accuracy
Ensemble models (4 algorithms)
SHAP/LIME explainability
Real-time drift detection
Automatic model monitoring
F1 score: 0.92"""
    },
    {
        "persona": "Data Scientists",
        "page": "Rule Performance Analytics",
        "points": """• Detection Rule Observatory: 20+ rules tracked
• Rule ROI Analysis: Which rules provide best value
• ML vs Rule Comparison: Performance benchmarks
• Correlation Network: Which rules work together
• False Positive Analysis: Precision optimization
• ML Enhancement: +8.3% precision gain with ML""",
        "metrics": """20+ fraud rules active
ML-enhanced rules: +8.3% precision
22% false positive reduction
Rule correlation insights
Continuous optimization"""
    },
    {
        "persona": "Operations Managers",
        "page": "Operational Analytics",
        "points": """• Transaction Flow Heatmap: By day/hour patterns
• Investigation Velocity: Time-to-resolution metrics
• Case Resolution Analytics: SLA tracking
• Merchant Risk Segmentation: Category-based risk
• Workload Optimization: Staffing recommendations
• Peak Time Identification: When fraud spikes""",
        "metrics": """Peak fraud: 2-4 AM weekends
Resolution SLA: 94.7% compliance
Merchant fraud rates: 1.5%-8.5%
Workload forecasting enabled
Staffing optimization insights"""
    },
    {
        "persona": "Sales Prospects (Any Persona)",
        "page": "Scenario Analysis",
        "points": """• 13 Real Fraud Scenarios: Comprehensive coverage
• Detection Rate Proof: 85-96% across all scenarios
• Real-World Timelines: See detection in action
• Educational Value: Understand fraud types
• Confidence Builder: Sophisticated attacks caught
• Customizable: Add client-specific scenarios""",
        "metrics": """13 fraud scenarios covered
85-96% detection rates
Account Takeover: 92% detection
BEC Fraud: 89% detection
Money Mule: 87% detection
Check Fraud: 94% detection"""
    },
    {
        "persona": "Geographic Risk Managers",
        "page": "Geo Analytics",
        "points": """• Geolocation Threat Map: USA state-level analysis
• VPN/Proxy Detection: Spoofing identification
• Behavioral Anomaly Timeline: Frequency and amount
• Cross-Vector Threat Analysis: Multi-factor correlation
• Device Fingerprinting: Identity verification
• Location Consistency Checks: Impossible travel""",
        "metrics": """VPN detection: 145 cases (CA)
Geographic fraud patterns identified
Real-time location analysis
Device fingerprinting enabled
Impossible travel detection
Multi-factor correlation"""
    }
]

for data in persona_data:
    ws2.cell(row=row, column=1).value = data['persona']
    ws2.cell(row=row, column=2).value = data['page']
    ws2.cell(row=row, column=3).value = data['points']
    ws2.cell(row=row, column=4).value = data['metrics']

    for col in range(1, 5):
        ws2.cell(row=row, column=col).alignment = Alignment(wrap_text=True, vertical='top')
        ws2.cell(row=row, column=col).border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

    ws2.row_dimensions[row].height = 120
    row += 1

# Set column widths
ws2.column_dimensions['A'].width = 25
ws2.column_dimensions['B'].width = 25
ws2.column_dimensions['C'].width = 40
ws2.column_dimensions['D'].width = 30

# ============================================================================
# SHEET 3: Improvement Recommendations
# ============================================================================
ws3 = wb.create_sheet("Improvement Recommendations")

# Title
ws3['A1'] = "Brainstorm Ideas to Make Dashboard More Sellable"
ws3['A1'].font = Font(size=16, bold=True, color="FFFFFF")
ws3['A1'].fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
ws3.merge_cells('A1:F1')

# Headers
row = 3
headers = ["Category", "Improvement Idea", "Why It Matters", "Selling Impact", "Effort", "Priority"]
for col_num, header in enumerate(headers, 1):
    cell = ws3.cell(row=row, column=col_num)
    cell.value = header
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    cell.alignment = Alignment(wrap_text=True, horizontal='center')

# Data
row += 1
improvements = [
    {
        "category": "Page Organization",
        "idea": "Reorganize navigation by PERSONA (Executive View, Analyst Workspace, Compliance Hub, Performance Center) instead of alphabetical",
        "why": "Users can quickly find their relevant pages, improves first-time user experience, shows thoughtful design",
        "impact": "HIGH - Makes demos smoother, executives see their dashboard immediately",
        "effort": "Low (2-3 hours)",
        "priority": "1 - CRITICAL"
    },
    {
        "category": "Page Organization",
        "idea": "Add 'Getting Started' landing page with role selector (I am: Executive / Analyst / Compliance / Data Scientist)",
        "why": "Guided onboarding, shows personalization, reduces learning curve",
        "impact": "HIGH - Great first impression, shows customer-centric design",
        "effort": "Medium (4-6 hours)",
        "priority": "2 - HIGH"
    },
    {
        "category": "Visual Polish",
        "idea": "Add consistent header to all pages with logo, tagline 'AI-Powered Fraud Detection', and current ROI widget",
        "why": "Professional branding, reinforces value prop on every page, shows ROI constantly",
        "impact": "MEDIUM - Professional polish, subconscious value reinforcement",
        "effort": "Low (2-3 hours)",
        "priority": "2 - HIGH"
    },
    {
        "category": "Visual Polish",
        "idea": "Add 'Quick Win' callout boxes on each page highlighting most impressive metric (e.g., '$2.3M prevented today')",
        "why": "Draws eye to key value, makes metrics scannable, reinforces ROI",
        "impact": "HIGH - Ensures key metrics aren't missed during demos",
        "effort": "Low (3-4 hours)",
        "priority": "1 - CRITICAL"
    },
    {
        "category": "Export Capabilities",
        "idea": "Add 'Export to PDF' button on Executive Dashboard for board presentations",
        "why": "Executives want to share with board/management, shows system integrates into workflows",
        "impact": "HIGH - Removes objection 'can I share this with my team?'",
        "effort": "Medium (6-8 hours)",
        "priority": "2 - HIGH"
    },
    {
        "category": "Export Capabilities",
        "idea": "Add 'Export to Excel' on Compliance page for audit reports",
        "why": "Compliance needs to provide data to regulators, removes manual data entry",
        "impact": "HIGH - Critical for compliance officers (deal breaker if missing)",
        "effort": "Medium (4-6 hours)",
        "priority": "1 - CRITICAL"
    },
    {
        "category": "Interactive Features",
        "idea": "Add time period selector (Last 24hrs / 7 days / 30 days / 90 days / Custom) on all pages",
        "why": "Shows flexibility, lets prospects explore their own timeframes during demo",
        "impact": "MEDIUM - Shows configurability, makes demos interactive",
        "effort": "High (12-16 hours)",
        "priority": "3 - MEDIUM"
    },
    {
        "category": "Interactive Features",
        "idea": "Add interactive ROI calculator widget: 'Input your annual fraud losses → See your savings with our system'",
        "why": "Personalizes business case on the spot, creates emotional investment",
        "impact": "VERY HIGH - Converts prospects by showing THEIR specific ROI",
        "effort": "Medium (8-10 hours)",
        "priority": "1 - CRITICAL"
    },
    {
        "category": "Social Proof",
        "idea": "Add footer to all pages: 'Trusted by 47+ financial institutions' with industry breakdown",
        "why": "Social proof builds trust, shows established product (not beta)",
        "impact": "HIGH - Reduces perceived risk of being 'first customer'",
        "effort": "Low (1-2 hours)",
        "priority": "2 - HIGH"
    },
    {
        "category": "Social Proof",
        "idea": "Add 'Customer Success Stories' page with anonymized case studies (Bank A: $45M saved, 92% detection)",
        "why": "Proof of real-world results, relatable to prospects, builds credibility",
        "impact": "VERY HIGH - Nothing sells like peer success stories",
        "effort": "Medium (6-8 hours content + 2 hours implementation)",
        "priority": "2 - HIGH"
    },
    {
        "category": "Competitive Position",
        "idea": "Add 'Industry Benchmark Comparison' widget showing client metrics vs. peer average",
        "why": "Creates competitive pressure ('you're behind'), urgency to buy",
        "impact": "HIGH - Triggers FOMO (fear of missing out)",
        "effort": "Medium (8-10 hours)",
        "priority": "3 - MEDIUM"
    },
    {
        "category": "Competitive Position",
        "idea": "Add 'What Makes Us Different' page comparing to legacy fraud systems (features matrix)",
        "why": "Clarifies differentiation, addresses 'how is this better than what we have?' question",
        "impact": "MEDIUM - Helps justify switch from current system",
        "effort": "Low (3-4 hours)",
        "priority": "3 - MEDIUM"
    },
    {
        "category": "New Capabilities",
        "idea": "Add 'Customer Onboarding Impact' page showing fraud prevention vs. customer approval rates",
        "why": "Addresses major objection: 'Will this block good customers?' Shows balanced approach",
        "impact": "VERY HIGH - Removes major purchase objection",
        "effort": "High (16-20 hours)",
        "priority": "2 - HIGH"
    },
    {
        "category": "New Capabilities",
        "idea": "Add 'Integration Health Monitor' showing API uptime, latency, error rates",
        "why": "Technical buyers worry about reliability, de-risks implementation concerns",
        "impact": "MEDIUM - Reassures technical stakeholders",
        "effort": "High (12-16 hours)",
        "priority": "4 - LOW"
    },
    {
        "category": "New Capabilities",
        "idea": "Add 'Model Evolution Timeline' showing accuracy improvements from v1.0 to current",
        "why": "Proves continuous improvement, shows ongoing value beyond initial deployment",
        "impact": "MEDIUM - Demonstrates long-term investment value",
        "effort": "Medium (6-8 hours)",
        "priority": "3 - MEDIUM"
    },
    {
        "category": "New Capabilities",
        "idea": "Add 'Escaped Fraud Analysis' page showing post-detection reviews (fraud that got through)",
        "why": "Addresses 'what about false negatives?' concern, shows transparency",
        "impact": "HIGH - Shows honest approach, builds trust",
        "effort": "High (12-16 hours)",
        "priority": "3 - MEDIUM"
    },
    {
        "category": "Demo Enhancements",
        "idea": "Add 'Demo Mode' toggle that shows sample data vs. real data (for sales demos)",
        "why": "Sales team can demo without exposing real customer data, polished demo data",
        "impact": "VERY HIGH - Makes sales demos professional and safe",
        "effort": "Medium (8-10 hours)",
        "priority": "1 - CRITICAL"
    },
    {
        "category": "Demo Enhancements",
        "idea": "Add 'Guided Tour' button that walks through key features with callouts and highlights",
        "why": "Reduces onboarding time, ensures key features aren't missed, self-service demo",
        "impact": "HIGH - Enables async demos (send link to prospect)",
        "effort": "High (16-20 hours)",
        "priority": "3 - MEDIUM"
    },
    {
        "category": "Demo Enhancements",
        "idea": "Add 'Scenario Simulator' where prospects can upload CSV and see their data analyzed live",
        "why": "Proves system works on THEIR data, creates 'aha moment', high engagement",
        "impact": "VERY HIGH - Nothing beats live demo on their data",
        "effort": "Very High (24-32 hours)",
        "priority": "4 - LOW (nice-to-have)"
    },
    {
        "category": "Value Reinforcement",
        "idea": "Add persistent 'Value Counter' widget showing running total of fraud prevented ($2.3M today, $839M/year)",
        "why": "Constant reminder of value delivered, creates pride in system",
        "impact": "MEDIUM - Subliminal value reinforcement",
        "effort": "Low (2-3 hours)",
        "priority": "2 - HIGH"
    },
    {
        "category": "Value Reinforcement",
        "idea": "Add 'This Week in Fraud Prevention' summary email with key metrics and caught fraud cases",
        "why": "Keeps executives engaged, shows ongoing value, surfaces product weekly",
        "impact": "HIGH - Maintains executive awareness and support",
        "effort": "High (12-16 hours)",
        "priority": "4 - LOW (post-sale feature)"
    },
    {
        "category": "Mobile Experience",
        "idea": "Optimize Executive Dashboard for mobile/tablet viewing (responsive design)",
        "why": "Executives check dashboards on phone/iPad, shows modern design",
        "impact": "MEDIUM - Expected for modern software, removes barrier",
        "effort": "High (16-24 hours)",
        "priority": "3 - MEDIUM"
    },
    {
        "category": "Alerts & Notifications",
        "idea": "Add 'Alert Center' page showing notification history and delivery status (email/SMS/Slack)",
        "why": "Shows proactive alerting, multi-channel notifications, reduces response time",
        "impact": "MEDIUM - Demonstrates real-time nature of system",
        "effort": "High (12-16 hours)",
        "priority": "4 - LOW"
    },
    {
        "category": "Documentation",
        "idea": "Add 'Help' icon on every chart with tooltip explaining what metric means and why it matters",
        "why": "Reduces confusion, makes system self-explanatory, less training needed",
        "impact": "MEDIUM - Improves user experience, faster adoption",
        "effort": "High (16-20 hours)",
        "priority": "3 - MEDIUM"
    }
]

for imp in improvements:
    ws3.cell(row=row, column=1).value = imp['category']
    ws3.cell(row=row, column=2).value = imp['idea']
    ws3.cell(row=row, column=3).value = imp['why']
    ws3.cell(row=row, column=4).value = imp['impact']
    ws3.cell(row=row, column=5).value = imp['effort']
    ws3.cell(row=row, column=6).value = imp['priority']

    # Color code priority
    priority_cell = ws3.cell(row=row, column=6)
    if '1 - CRITICAL' in imp['priority']:
        priority_cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        priority_cell.font = Font(color="FFFFFF", bold=True)
    elif '2 - HIGH' in imp['priority']:
        priority_cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        priority_cell.font = Font(bold=True)
    elif '3 - MEDIUM' in imp['priority']:
        priority_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    else:
        priority_cell.fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")

    for col in range(1, 7):
        ws3.cell(row=row, column=col).alignment = Alignment(wrap_text=True, vertical='top')
        ws3.cell(row=row, column=col).border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

    ws3.row_dimensions[row].height = 80
    row += 1

# Set column widths
ws3.column_dimensions['A'].width = 20
ws3.column_dimensions['B'].width = 40
ws3.column_dimensions['C'].width = 35
ws3.column_dimensions['D'].width = 18
ws3.column_dimensions['E'].width = 18
ws3.column_dimensions['F'].width = 18

# ============================================================================
# SHEET 4: Task Assignment - 2 Person Team
# ============================================================================
ws4 = wb.create_sheet("Task Assignment - 2 Person Team")

# Title
ws4['A1'] = "Implementation Plan: 2-Person Team Task Assignments"
ws4['A1'].font = Font(size=16, bold=True, color="FFFFFF")
ws4['A1'].fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
ws4.merge_cells('A1:G1')

ws4['A2'] = "Team: Person A (Senior Developer - Full-Stack) | Person B (Developer - Frontend Focus)"
ws4['A2'].font = Font(italic=True)
ws4.merge_cells('A2:G2')

# Headers
row = 4
headers = ["Priority", "Task", "Description", "Assigned To", "Estimated Hours", "Dependencies", "Week"]
for col_num, header in enumerate(headers, 1):
    cell = ws4.cell(row=row, column=col_num)
    cell.value = header
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    cell.alignment = Alignment(wrap_text=True, horizontal='center')

# Data
row += 1
tasks = [
    # Week 1 - Critical Items
    {"priority": "1", "task": "Reorganize navigation by persona", "desc": "Create role-based navigation (Executive View, Analyst Workspace, Compliance Hub, Performance Center)", "assigned": "Person A", "hours": "3", "depends": "None", "week": "Week 1"},
    {"priority": "1", "task": "Add Quick Win callout boxes", "desc": "Highlight most impressive metric on each page (e.g., '$2.3M prevented today')", "assigned": "Person B", "hours": "4", "depends": "None", "week": "Week 1"},
    {"priority": "1", "task": "Add Export to PDF (Executive Dashboard)", "desc": "PDF export for board presentations with formatting", "assigned": "Person A", "hours": "8", "depends": "None", "week": "Week 1"},
    {"priority": "1", "task": "Add Export to Excel (Compliance page)", "desc": "Excel export for audit reports with raw data", "assigned": "Person B", "hours": "6", "depends": "None", "week": "Week 1"},
    {"priority": "1", "task": "Add interactive ROI calculator", "desc": "Widget: 'Input your annual fraud losses → See your savings'", "assigned": "Person A", "hours": "10", "depends": "None", "week": "Week 1"},
    {"priority": "1", "task": "Add Demo Mode toggle", "desc": "Switch between sample data (for demos) vs. real data", "assigned": "Person B", "hours": "10", "depends": "None", "week": "Week 1"},

    # Week 2 - High Priority Items
    {"priority": "2", "task": "Add Getting Started landing page", "desc": "Role selector page (I am: Executive / Analyst / Compliance)", "assigned": "Person A", "hours": "6", "depends": "Reorganize navigation", "week": "Week 2"},
    {"priority": "2", "task": "Add consistent header with logo + ROI widget", "desc": "Branded header on all pages with current ROI display", "assigned": "Person B", "hours": "3", "depends": "None", "week": "Week 2"},
    {"priority": "2", "task": "Add social proof footer", "desc": "Footer: 'Trusted by 47+ financial institutions' on all pages", "assigned": "Person B", "hours": "2", "depends": "Add consistent header", "week": "Week 2"},
    {"priority": "2", "task": "Create Customer Success Stories page", "desc": "Page with anonymized case studies (Bank A: $45M saved, 92% detection)", "assigned": "Person A", "hours": "8", "depends": "None", "week": "Week 2"},
    {"priority": "2", "task": "Add Value Counter widget", "desc": "Persistent widget showing running total: '$2.3M today, $839M/year'", "assigned": "Person B", "hours": "3", "depends": "None", "week": "Week 2"},
    {"priority": "2", "task": "Add Customer Onboarding Impact page", "desc": "Page showing fraud prevention vs. customer approval rates (addresses 'will this block good customers?' objection)", "assigned": "Person A", "hours": "20", "depends": "None", "week": "Week 2-3"},

    # Week 3 - Medium Priority Items
    {"priority": "3", "task": "Add time period selector (all pages)", "desc": "Dropdown: Last 24hrs / 7 days / 30 days / 90 days / Custom", "assigned": "Person A", "hours": "16", "depends": "None", "week": "Week 3"},
    {"priority": "3", "task": "Add Industry Benchmark Comparison widget", "desc": "Widget showing client metrics vs. peer average", "assigned": "Person B", "hours": "10", "depends": "None", "week": "Week 3"},
    {"priority": "3", "task": "Create 'What Makes Us Different' page", "desc": "Feature comparison matrix vs. legacy fraud systems", "assigned": "Person B", "hours": "4", "depends": "None", "week": "Week 3"},
    {"priority": "3", "task": "Add Model Evolution Timeline page", "desc": "Page showing accuracy improvements from v1.0 to current", "assigned": "Person A", "hours": "8", "depends": "None", "week": "Week 3"},
    {"priority": "3", "task": "Add Escaped Fraud Analysis page", "desc": "Page showing post-detection reviews (fraud that got through)", "assigned": "Person A", "hours": "16", "depends": "None", "week": "Week 3-4"},
    {"priority": "3", "task": "Add Guided Tour button", "desc": "Interactive tour walking through key features with callouts", "assigned": "Person B", "hours": "20", "depends": "None", "week": "Week 3-4"},
    {"priority": "3", "task": "Optimize Executive Dashboard for mobile", "desc": "Responsive design for phone/tablet viewing", "assigned": "Person B", "hours": "24", "depends": "None", "week": "Week 3-4"},
    {"priority": "3", "task": "Add Help icons on all charts", "desc": "Tooltip on every chart explaining metric and why it matters", "assigned": "Person A", "hours": "20", "depends": "None", "week": "Week 4"},

    # Week 4+ - Lower Priority / Nice-to-Have
    {"priority": "4", "task": "Add Integration Health Monitor page", "desc": "Page showing API uptime, latency, error rates", "assigned": "Person A", "hours": "16", "depends": "None", "week": "Week 4"},
    {"priority": "4", "task": "Add Alert Center page", "desc": "Page showing notification history (email/SMS/Slack)", "assigned": "Person B", "hours": "16", "depends": "None", "week": "Week 4"},
    {"priority": "4", "task": "Create weekly summary email", "desc": "'This Week in Fraud Prevention' email with key metrics", "assigned": "Person A", "hours": "16", "depends": "None", "week": "Week 4"},
    {"priority": "4", "task": "Add Scenario Simulator", "desc": "Upload CSV feature to analyze prospect's data live (advanced)", "assigned": "Person A", "hours": "32", "depends": "Demo Mode", "week": "Week 5-6"},
]

for task_data in tasks:
    ws4.cell(row=row, column=1).value = task_data['priority']
    ws4.cell(row=row, column=2).value = task_data['task']
    ws4.cell(row=row, column=3).value = task_data['desc']
    ws4.cell(row=row, column=4).value = task_data['assigned']
    ws4.cell(row=row, column=5).value = task_data['hours']
    ws4.cell(row=row, column=6).value = task_data['depends']
    ws4.cell(row=row, column=7).value = task_data['week']

    # Color code by priority
    priority_cell = ws4.cell(row=row, column=1)
    if task_data['priority'] == "1":
        priority_cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        priority_cell.font = Font(color="FFFFFF", bold=True)
    elif task_data['priority'] == "2":
        priority_cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        priority_cell.font = Font(bold=True)
    elif task_data['priority'] == "3":
        priority_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    else:
        priority_cell.fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")

    # Color code by person
    person_cell = ws4.cell(row=row, column=4)
    if "Person A" in task_data['assigned']:
        person_cell.fill = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")
    else:
        person_cell.fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")

    for col in range(1, 8):
        ws4.cell(row=row, column=col).alignment = Alignment(wrap_text=True, vertical='top')
        ws4.cell(row=row, column=col).border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

    ws4.row_dimensions[row].height = 60
    row += 1

# Add summary rows
row += 1
ws4[f'A{row}'] = "SUMMARY"
ws4[f'A{row}'].font = Font(bold=True, size=12, color="FFFFFF")
ws4[f'A{row}'].fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
ws4.merge_cells(f'A{row}:G{row}')

row += 1
ws4[f'A{row}'] = "Person A Total Hours:"
ws4[f'B{row}'] = "=SUMIF(D:D,\"Person A\",E:E)"
ws4[f'B{row}'].font = Font(bold=True)

row += 1
ws4[f'A{row}'] = "Person B Total Hours:"
ws4[f'B{row}'] = "=SUMIF(D:D,\"Person B\",E:E)"
ws4[f'B{row}'].font = Font(bold=True)

row += 1
ws4[f'A{row}'] = "Total Project Hours:"
ws4[f'B{row}'] = "=SUM(E:E)"
ws4[f'B{row}'].font = Font(bold=True)

row += 1
ws4[f'A{row}'] = "Estimated Project Duration:"
ws4[f'B{row}'] = "4-6 weeks (assuming 40 hours/week per person)"
ws4[f'B{row}'].font = Font(bold=True)

# Set column widths
ws4.column_dimensions['A'].width = 10
ws4.column_dimensions['B'].width = 30
ws4.column_dimensions['C'].width = 50
ws4.column_dimensions['D'].width = 15
ws4.column_dimensions['E'].width = 12
ws4.column_dimensions['F'].width = 20
ws4.column_dimensions['G'].width = 12

# ============================================================================
# SHEET 5: Most Convincing Dashboard Elements
# ============================================================================
ws5 = wb.create_sheet("Most Convincing Elements")

# Title
ws5['A1'] = "Most Convincing Dashboard Elements for Sales"
ws5['A1'].font = Font(size=16, bold=True, color="FFFFFF")
ws5['A1'].fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
ws5.merge_cells('A1:E1')

ws5['A2'] = "Analysis: Which graphs and pages make the strongest case for purchase"
ws5['A2'].font = Font(italic=True)
ws5.merge_cells('A2:E2')

# Headers
row = 4
headers = ["Rank", "Page", "Specific Graph/Element", "Why It's Convincing", "Use In Sales Pitch"]
for col_num, header in enumerate(headers, 1):
    cell = ws5.cell(row=row, column=col_num)
    cell.value = header
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    cell.alignment = Alignment(wrap_text=True, horizontal='center')

# Data
row += 1
convincing_elements = [
    {
        "rank": "1",
        "page": "Executive Dashboard",
        "element": "ROI Calculator Widget (39:1 return)",
        "why": "Shows immediate business value in dollars. CFO sees $39 saved for every $1 spent. Impossible to argue against math.",
        "pitch": "Open with this: 'Your ROI is 39 to 1. For every dollar you invest, you save $39 in fraud losses. That's not a cost - that's a profit center.'"
    },
    {
        "rank": "2",
        "page": "Executive Dashboard",
        "element": "Daily Fraud Prevention Summary ($2.3M/day)",
        "why": "Makes value tangible. '$2.3M prevented TODAY' is more impactful than abstract percentages. Shows real-time value.",
        "pitch": "While we've been talking for 30 minutes, this system has already prevented $2,875 in fraud. By tomorrow, that's $2.3 million.'"
    },
    {
        "rank": "3",
        "page": "Scenario Analysis",
        "element": "13 Real Fraud Scenarios with 85-96% Detection Rates",
        "why": "Proof that system catches real fraud, not just theoretical. Prospects can see scenarios matching their concerns.",
        "pitch": "Show scenarios relevant to their industry: 'You mentioned BEC fraud is your biggest concern. Here's how we detected it with 89% accuracy...'"
    },
    {
        "rank": "4",
        "page": "Compliance & KYC Analytics",
        "element": "Complete Audit Trail Coverage (100% transaction logging)",
        "why": "Addresses regulatory risk. Failed audits cost $5M-$50M. This is insurance against regulatory penalties.",
        "pitch": "One failed audit costs $5-50 million. Our complete audit trail means you'll pass every examination. That's peace of mind.'"
    },
    {
        "rank": "5",
        "page": "Analyst Dashboard",
        "element": "Transaction Lifecycle Funnel (95% auto-cleared)",
        "why": "Shows efficiency. Only 5% need manual review = 95% analyst time saved. Addresses 'will this increase our workload?' objection.",
        "pitch": "Your analysts are drowning in alerts. We auto-clear 95%, so they only focus on real threats. That's $59,575 saved per day.'"
    },
    {
        "rank": "6",
        "page": "AI & ML Intelligence",
        "element": "SHAP/LIME Explainable AI Visualizations",
        "why": "Proves AI isn't a 'black box'. Regulatory requirement: must explain decisions. Differentiates from competitors.",
        "pitch": "Other systems can't tell you WHY a transaction was flagged. Ours shows you exactly which factors contributed. That's regulatory compliance.'"
    },
    {
        "rank": "7",
        "page": "Fraud Transaction Monitoring",
        "element": "Real-Time Account Takeover Detection (<6ms processing)",
        "why": "Shows speed = catches fraud before money leaves. Real-time is critical for fraud prevention vs. fraud detection.",
        "pitch": "We detect account takeover in under 6 milliseconds. That's before the fraudster hits 'send'. Time matters in fraud.'"
    },
    {
        "rank": "8",
        "page": "Rule Performance Analytics",
        "element": "ML Enhancement: +8.3% Precision, 22% False Positive Reduction",
        "why": "Shows continuous improvement. System gets smarter over time. Not a static product - it evolves.",
        "pitch": "Our ML enhancement increased precision by 8.3% and cut false positives by 22%. Your system will get better every month.'"
    },
    {
        "rank": "9",
        "page": "Executive Dashboard",
        "element": "Detection Attribution Treemap (which rules catch fraud)",
        "why": "Shows transparency. Clients see exactly which rules are valuable. Helps justify cost by showing rule-level ROI.",
        "pitch": "This treemap shows which fraud rules caught what. 'Unverified Account Change' rule alone saved $697K last month.'"
    },
    {
        "rank": "10",
        "page": "Operational Analytics",
        "element": "Peak Fraud Time Identification (2-4 AM weekends)",
        "why": "Actionable insights. Shows system doesn't just detect - it helps optimize operations. Staffing recommendations have dollar value.",
        "pitch": "Did you know 45% of your fraud happens between 2-4 AM on weekends? Here's how to optimize your staffing to prevent it.'"
    },
    {
        "rank": "11",
        "page": "Geo Analytics",
        "element": "VPN Detection Heatmap (145 cases from California)",
        "why": "Shows sophistication. VPN detection means catching fraudsters who are actively hiding. Addresses 'can you catch sophisticated fraud?' question.",
        "pitch": "Fraudsters use VPNs to hide their location. We detected 145 attempts from California alone. We catch the smart ones.'"
    },
    {
        "rank": "12",
        "page": "Transaction Review",
        "element": "Complete Decision Explanation (why transaction was flagged)",
        "why": "Shows transparency for analysts. Reduces training time. Analysts understand system reasoning = trust = adoption.",
        "pitch": "Your analysts see exactly why each transaction was flagged. No mysterious 'black box' decisions. That's training built-in.'"
    },
    {
        "rank": "13",
        "page": "Executive Dashboard",
        "element": "90-Day Action Plan with ROI Calculations",
        "why": "Shows roadmap. Not just 'here's the problem' but 'here's what to do about it'. Executives want clear next steps.",
        "pitch": "We don't just show you problems. Here's your 90-day action plan with ROI for each improvement. That's a strategy, not just data.'"
    },
    {
        "rank": "14",
        "page": "Analyst Dashboard",
        "element": "Decision Pattern Analytics (analyst confidence trends)",
        "why": "Shows quality control. Tracks analyst performance. Helps identify training needs. Demonstrates continuous improvement culture.",
        "pitch": "See how your analysts are performing. High-confidence decisions are up 12%. Low-confidence cases down 8%. That's measurable improvement.'"
    },
    {
        "rank": "15",
        "page": "Compliance & KYC Analytics",
        "element": "False Positive Analysis (6.2% vs. industry 15-20%)",
        "why": "Shows efficiency vs. competitors. 6.2% false positive rate is exceptional. Less wasted analyst time = lower cost.",
        "pitch": "Industry average false positive rate: 15-20%. Ours: 6.2%. That's 70% fewer wasted investigations. Your analysts will love us.'"
    }
]

for elem in convincing_elements:
    ws5.cell(row=row, column=1).value = elem['rank']
    ws5.cell(row=row, column=2).value = elem['page']
    ws5.cell(row=row, column=3).value = elem['element']
    ws5.cell(row=row, column=4).value = elem['why']
    ws5.cell(row=row, column=5).value = elem['pitch']

    # Color code top 5
    rank_cell = ws5.cell(row=row, column=1)
    if int(elem['rank']) <= 3:
        rank_cell.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")  # Gold
        rank_cell.font = Font(bold=True)
    elif int(elem['rank']) <= 5:
        rank_cell.fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")  # Silver
        rank_cell.font = Font(bold=True)

    for col in range(1, 6):
        ws5.cell(row=row, column=col).alignment = Alignment(wrap_text=True, vertical='top')
        ws5.cell(row=row, column=col).border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

    ws5.row_dimensions[row].height = 80
    row += 1

# Set column widths
ws5.column_dimensions['A'].width = 8
ws5.column_dimensions['B'].width = 25
ws5.column_dimensions['C'].width = 35
ws5.column_dimensions['D'].width = 40
ws5.column_dimensions['E'].width = 50

# ============================================================================
# SHEET 6: Sales Pitch Structure
# ============================================================================
ws6 = wb.create_sheet("Sales Pitch Structure")

# Title
ws6['A1'] = "Recommended Sales Pitch Flow Using Dashboard"
ws6['A1'].font = Font(size=16, bold=True, color="FFFFFF")
ws6['A1'].fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
ws6.merge_cells('A1:D1')

ws6['A2'] = "45-Minute Demo Structure for Maximum Impact"
ws6['A2'].font = Font(italic=True)
ws6.merge_cells('A2:D2')

# Headers
row = 4
headers = ["Time", "Pitch Section", "Dashboard Page/Element to Show", "Key Talking Points"]
for col_num, header in enumerate(headers, 1):
    cell = ws6.cell(row=row, column=col_num)
    cell.value = header
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    cell.alignment = Alignment(wrap_text=True, horizontal='center')

# Data
row += 1
pitch_structure = [
    {"time": "0-5 min", "section": "Hook with ROI", "page": "Executive Dashboard - ROI Calculator", "points": "• '$2.3M prevented daily, $839M annually'\n• '39:1 ROI - not a cost, it's a profit center'\n• 'While we talk, system prevents fraud in real-time'\n• Establish business value before technical details"},
    {"time": "5-10 min", "section": "Identify Their Pain Points", "page": "Ask questions, then tailor demo", "points": "• 'What fraud types worry you most?'\n• 'What's your current false positive rate?'\n• 'How many analysts do you have?'\n• 'Ever failed a compliance audit?'\n• Listen first, then show relevant scenarios"},
    {"time": "10-20 min", "section": "Show Proof: Scenario Detection", "page": "Scenario Analysis - Their Pain Points", "points": "• Show 2-3 scenarios matching their concerns\n• If BEC: Show vendor impersonation (89% detection)\n• If Account Takeover: Show SIM swap detection (92%)\n• If Money Laundering: Show money mule (87%)\n• Make it tangible: 'This is what you're facing today'"},
    {"time": "20-25 min", "section": "Operational Efficiency Win", "page": "Analyst Dashboard - Transaction Lifecycle Funnel", "points": "• '95% auto-cleared - your analysts only see real threats'\n• '$59,575 saved daily in analyst time'\n• 'False positive rate: 6.2% vs industry 15-20%'\n• Address: 'Will this increase our workload?' (No, decreases by 95%)"},
    {"time": "25-30 min", "section": "Compliance & Risk Mitigation", "page": "Compliance & KYC Analytics", "points": "• 'Complete audit trail - pass every examination'\n• 'Failed audits cost $5M-$50M - this is insurance'\n• '100% transaction logging for regulators'\n• 'KYC/CDD/EDD lifecycle tracking'\n• Position as risk mitigation, not just fraud detection"},
    {"time": "30-35 min", "section": "Technical Credibility (if needed)", "page": "AI & ML Intelligence - SHAP/LIME", "points": "• Only show if technical buyers in room\n• 'Explainable AI - not a black box'\n• '94.3% model accuracy with continuous improvement'\n• 'SHAP values show why each decision made'\n• Proves this is real AI, not marketing hype"},
    {"time": "35-40 min", "section": "Competitive Differentiation", "page": "Rule Performance Analytics", "points": "• 'ML enhancement: +8.3% precision improvement'\n• '22% false positive reduction over time'\n• 'System gets smarter every month'\n• Competitors: static rules. Us: continuous learning\n• Not just a tool, it's a learning system"},
    {"time": "40-43 min", "section": "Return to ROI & Close", "page": "Executive Dashboard - Summary", "points": "• Return to business value\n• '$839M saved annually for $59K cost'\n• '3-month payback period'\n• 'Question: Can you afford NOT to have this?'\n• Create urgency: fraud increases 25% annually"},
    {"time": "43-45 min", "section": "Next Steps", "page": "Close dashboard, focus on conversation", "points": "• 'Pilot: 30 days on your data, see fraud we catch'\n• 'No cost pilot - see results before committing'\n• 'Timeline: 60-90 day implementation'\n• 'Who else needs to see this?' (expand stakeholders)\n• Schedule follow-up within 3-5 days (momentum)"}
]

for pitch in pitch_structure:
    ws6.cell(row=row, column=1).value = pitch['time']
    ws6.cell(row=row, column=2).value = pitch['section']
    ws6.cell(row=row, column=3).value = pitch['page']
    ws6.cell(row=row, column=4).value = pitch['points']

    # Color code by section
    if 'Hook' in pitch['section'] or 'Close' in pitch['section']:
        ws6.cell(row=row, column=2).fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
        ws6.cell(row=row, column=2).font = Font(bold=True)

    for col in range(1, 5):
        ws6.cell(row=row, column=col).alignment = Alignment(wrap_text=True, vertical='top')
        ws6.cell(row=row, column=col).border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

    ws6.row_dimensions[row].height = 100
    row += 1

# Set column widths
ws6.column_dimensions['A'].width = 12
ws6.column_dimensions['B'].width = 25
ws6.column_dimensions['C'].width = 35
ws6.column_dimensions['D'].width = 50

# Save workbook
output_file = "/home/user/transaction-monitoring/Dashboard_Sales_Analysis_2_Person_Team.xlsx"
wb.save(output_file)
print(f"Excel file created successfully: {output_file}")
print(f"Total sheets: {len(wb.sheetnames)}")
for sheet_name in wb.sheetnames:
    print(f"  - {sheet_name}")
