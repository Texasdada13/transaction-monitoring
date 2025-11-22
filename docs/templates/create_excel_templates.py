#!/usr/bin/env python3
"""
Script to generate Excel templates for the Application Portfolio Rationalization project.
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime, timedelta

# Ensure output directory exists
output_dir = os.path.dirname(os.path.abspath(__file__))


def create_project_plan_detailed():
    """Create PROJECT_PLAN_DETAILED.xlsx with multiple sheets."""
    wb = Workbook()

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="002B5B", end_color="002B5B", fill_type="solid")
    subheader_fill = PatternFill(start_color="0A5CAD", end_color="0A5CAD", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Sheet 1: Project Overview
    ws = wb.active
    ws.title = "Project Overview"

    overview_data = [
        ["APPLICATION PORTFOLIO RATIONALIZATION", "", "", ""],
        ["PROJECT PLAN", "", "", ""],
        ["", "", "", ""],
        ["Project Information", "", "", ""],
        ["Project Name:", "[Client Name] APR Engagement", "", ""],
        ["Project Manager:", "[PM Name]", "", ""],
        ["Start Date:", datetime.now().strftime("%Y-%m-%d"), "", ""],
        ["End Date:", (datetime.now() + timedelta(weeks=10)).strftime("%Y-%m-%d"), "", ""],
        ["Duration:", "10 weeks", "", ""],
        ["", "", "", ""],
        ["Phase Summary", "Start", "End", "Duration"],
        ["Phase 1: Initiation", "Week 1", "Week 1", "1 week"],
        ["Phase 2: Discovery", "Week 2", "Week 3", "2 weeks"],
        ["Phase 3: Assessment", "Week 3", "Week 5", "3 weeks"],
        ["Phase 4: Analysis", "Week 5", "Week 6", "2 weeks"],
        ["Phase 5: Recommendations", "Week 6", "Week 8", "3 weeks"],
        ["Phase 6: Validation", "Week 8", "Week 9", "2 weeks"],
        ["Phase 7: Delivery", "Week 9", "Week 10", "2 weeks"],
    ]

    for row_idx, row in enumerate(overview_data, 1):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if row_idx in [1, 2]:
                cell.font = Font(bold=True, size=16)
            elif row_idx in [4, 11]:
                cell.font = Font(bold=True)
                cell.fill = subheader_fill
                cell.font = header_font
            elif row_idx >= 12:
                cell.border = thin_border

    # Adjust column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15

    # Sheet 2: Detailed Tasks
    ws2 = wb.create_sheet("Detailed Tasks")

    task_headers = ["Task ID", "Phase", "Task Name", "Description", "Owner", "Start Date",
                   "End Date", "Duration (days)", "Dependencies", "Status", "% Complete", "Notes"]

    for col, header in enumerate(task_headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # Sample tasks (45+ items)
    tasks = [
        ["1.1", "Initiation", "Conduct kickoff meeting", "Present project plan and objectives", "PM", "", "", "1", "", "Not Started", "0%", ""],
        ["1.2", "Initiation", "Identify stakeholders", "Create stakeholder registry", "BA", "", "", "2", "1.1", "Not Started", "0%", ""],
        ["1.3", "Initiation", "Inventory data sources", "Document all available data sources", "BA", "", "", "3", "1.1", "Not Started", "0%", ""],
        ["1.4", "Initiation", "Setup project tools", "Configure assessment tool access", "PM", "", "", "3", "1.1", "Not Started", "0%", ""],
        ["1.5", "Initiation", "Create communication plan", "Define status reporting cadence", "PM", "", "", "1", "1.1", "Not Started", "0%", ""],
        ["2.1", "Discovery", "Extract CMDB data", "Pull application inventory from CMDB", "IT Ops", "", "", "3", "1.3", "Not Started", "0%", ""],
        ["2.2", "Discovery", "Collect financial data", "Gather cost data by application", "Finance", "", "", "5", "1.3", "Not Started", "0%", ""],
        ["2.3", "Discovery", "Gather usage metrics", "Extract APM and usage data", "Architect", "", "", "3", "2.1", "Not Started", "0%", ""],
        ["2.4", "Discovery", "Map integrations", "Document application dependencies", "Architect", "", "", "3", "2.1", "Not Started", "0%", ""],
        ["2.5", "Discovery", "Validate data quality", "Assess completeness and accuracy", "BA", "", "", "3", "2.2,2.3", "Not Started", "0%", ""],
        ["2.6", "Discovery", "Document data gaps", "Identify missing information", "BA", "", "", "1", "2.5", "Not Started", "0%", ""],
        ["3.1", "Assessment", "Customize questionnaire", "Adapt questions for client context", "BA", "", "", "1", "2.5", "Not Started", "0%", ""],
        ["3.2", "Assessment", "Schedule interviews", "Coordinate with stakeholders", "PM", "", "", "2", "1.2", "Not Started", "0%", ""],
        ["3.3", "Assessment", "Conduct interviews - Week 1", "Complete first batch of interviews", "BA", "", "", "5", "3.1,3.2", "Not Started", "0%", ""],
        ["3.4", "Assessment", "Conduct interviews - Week 2", "Complete second batch of interviews", "BA", "", "", "5", "3.3", "Not Started", "0%", ""],
        ["3.5", "Assessment", "Compile responses", "Enter all responses into tool", "BA", "", "", "2", "3.4", "Not Started", "0%", ""],
        ["3.6", "Assessment", "Calculate assessment scores", "Generate scores by category", "BA", "", "", "1", "3.5", "Not Started", "0%", ""],
        ["3.7", "Assessment", "Identify gaps", "Note applications without assessment", "BA", "", "", "1", "3.6", "Not Started", "0%", ""],
        ["4.1", "Analysis", "Apply quantitative scoring", "Score technical and cost metrics", "Architect", "", "", "3", "2.5", "Not Started", "0%", ""],
        ["4.2", "Analysis", "Integrate qualitative data", "Combine with assessment scores", "BA", "", "", "2", "3.6,4.1", "Not Started", "0%", ""],
        ["4.3", "Analysis", "Apply TIME framework", "Assign disposition to each app", "Lead", "", "", "2", "4.2", "Not Started", "0%", ""],
        ["4.4", "Analysis", "Perform cluster analysis", "Group similar applications", "Architect", "", "", "2", "4.3", "Not Started", "0%", ""],
        ["4.5", "Analysis", "Identify opportunities", "List potential savings/improvements", "Lead", "", "", "2", "4.4", "Not Started", "0%", ""],
        ["5.1", "Recommendations", "Develop retire recommendations", "Detail decommission approach", "Lead", "", "", "2", "4.5", "Not Started", "0%", ""],
        ["5.2", "Recommendations", "Develop consolidate recommendations", "Identify merge opportunities", "Architect", "", "", "2", "4.5", "Not Started", "0%", ""],
        ["5.3", "Recommendations", "Develop modernize recommendations", "Plan enhancement initiatives", "Architect", "", "", "2", "4.5", "Not Started", "0%", ""],
        ["5.4", "Recommendations", "Develop replace recommendations", "Identify SaaS alternatives", "BA", "", "", "2", "4.5", "Not Started", "0%", ""],
        ["5.5", "Recommendations", "Create implementation roadmap", "Sequence all initiatives", "Lead", "", "", "3", "5.1-5.4", "Not Started", "0%", ""],
        ["5.6", "Recommendations", "Develop business case", "Calculate ROI and payback", "BA", "", "", "3", "5.5", "Not Started", "0%", ""],
        ["5.7", "Recommendations", "Assess risks", "Document risks and mitigations", "Architect", "", "", "2", "5.5", "Not Started", "0%", ""],
        ["5.8", "Recommendations", "Identify quick wins", "List immediate opportunities", "Lead", "", "", "1", "5.5", "Not Started", "0%", ""],
        ["6.1", "Validation", "Conduct IT validation session", "Review with technical team", "Lead", "", "", "1", "5.6", "Not Started", "0%", ""],
        ["6.2", "Validation", "Conduct business validation", "Review with business owners", "Lead", "", "", "1", "6.1", "Not Started", "0%", ""],
        ["6.3", "Validation", "Incorporate feedback", "Refine recommendations", "Team", "", "", "3", "6.2", "Not Started", "0%", ""],
        ["6.4", "Validation", "Prepare executive materials", "Create presentation deck", "Lead", "", "", "2", "6.3", "Not Started", "0%", ""],
        ["6.5", "Validation", "Compile final report", "Assemble all deliverables", "BA", "", "", "2", "6.3", "Not Started", "0%", ""],
        ["7.1", "Delivery", "Deliver executive presentation", "Present to leadership", "Lead", "", "", "1", "6.4", "Not Started", "0%", ""],
        ["7.2", "Delivery", "Conduct tool training", "Train client team on tool", "Architect", "", "", "1", "6.5", "Not Started", "0%", ""],
        ["7.3", "Delivery", "Document processes", "Create process guides", "BA", "", "", "2", "7.1", "Not Started", "0%", ""],
        ["7.4", "Delivery", "Hand off data", "Transfer all project data", "PM", "", "", "1", "7.2", "Not Started", "0%", ""],
        ["7.5", "Delivery", "Conduct closeout meeting", "Formal project closure", "Lead", "", "", "1", "7.4", "Not Started", "0%", ""],
        ["7.6", "Delivery", "Collect feedback", "Client satisfaction survey", "PM", "", "", "1", "7.5", "Not Started", "0%", ""],
    ]

    for row_idx, task in enumerate(tasks, 2):
        for col_idx, value in enumerate(task, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border

    # Add data validation for Status
    status_validation = DataValidation(
        type="list",
        formula1='"Not Started,In Progress,Completed,On Hold,Cancelled"',
        allow_blank=True
    )
    ws2.add_data_validation(status_validation)
    status_validation.add(f'J2:J{len(tasks)+1}')

    # Adjust column widths
    col_widths = [8, 15, 30, 40, 12, 12, 12, 12, 15, 12, 10, 30]
    for i, width in enumerate(col_widths, 1):
        ws2.column_dimensions[get_column_letter(i)].width = width

    # Sheet 3: Stakeholder Registry
    ws3 = wb.create_sheet("Stakeholder Registry")

    stakeholder_headers = ["ID", "Name", "Title", "Department", "Email", "Phone",
                          "Stakeholder Type", "Applications", "Influence", "Interest",
                          "Communication Preference", "Notes"]

    for col, header in enumerate(stakeholder_headers, 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill

    # Sample stakeholders
    stakeholders = [
        ["S001", "John Smith", "VP Operations", "Operations", "john.smith@client.com", "555-0101", "Executive", "ERP, CRM", "High", "High", "Email", "Executive sponsor"],
        ["S002", "Sarah Johnson", "IT Director", "IT", "sarah.johnson@client.com", "555-0102", "Technical Lead", "All", "High", "High", "Teams", "Key technical contact"],
        ["S003", "Mike Chen", "Finance Manager", "Finance", "mike.chen@client.com", "555-0103", "Business Owner", "Finance Apps", "Medium", "High", "Email", "Cost data owner"],
    ]

    for row_idx, stakeholder in enumerate(stakeholders, 2):
        for col_idx, value in enumerate(stakeholder, 1):
            cell = ws3.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border

    # Adjust column widths
    col_widths = [6, 20, 20, 15, 30, 12, 15, 20, 10, 10, 20, 30]
    for i, width in enumerate(col_widths, 1):
        ws3.column_dimensions[get_column_letter(i)].width = width

    # Sheet 4: Deliverables Tracker
    ws4 = wb.create_sheet("Deliverables Tracker")

    deliverable_headers = ["ID", "Deliverable Name", "Phase", "Due Date", "Owner",
                          "Reviewers", "Status", "Completion Date", "Location", "Notes"]

    for col, header in enumerate(deliverable_headers, 1):
        cell = ws4.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill

    deliverables = [
        ["D001", "Project Charter", "1", "", "PM", "Sponsor", "Not Started", "", "", ""],
        ["D002", "Stakeholder Registry", "1", "", "PM", "Lead", "Not Started", "", "", ""],
        ["D003", "Application Inventory", "2", "", "BA", "Architect", "Not Started", "", "", ""],
        ["D004", "Data Quality Report", "2", "", "BA", "Lead", "Not Started", "", "", ""],
        ["D005", "Assessment Scores", "3", "", "BA", "Lead", "Not Started", "", "", ""],
        ["D006", "TIME Disposition Matrix", "4", "", "Lead", "Team", "Not Started", "", "", ""],
        ["D007", "Recommendation Report", "5", "", "Lead", "Architect", "Not Started", "", "", ""],
        ["D008", "Implementation Roadmap", "5", "", "Architect", "Lead", "Not Started", "", "", ""],
        ["D009", "Business Case", "5", "", "BA", "Finance", "Not Started", "", "", ""],
        ["D010", "Executive Presentation", "6", "", "Lead", "Sponsor", "Not Started", "", "", ""],
        ["D011", "Final Report", "6", "", "BA", "Lead", "Not Started", "", "", ""],
        ["D012", "Training Materials", "7", "", "Architect", "PM", "Not Started", "", "", ""],
    ]

    for row_idx, deliverable in enumerate(deliverables, 2):
        for col_idx, value in enumerate(deliverable, 1):
            cell = ws4.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border

    col_widths = [6, 25, 8, 12, 12, 15, 12, 15, 25, 30]
    for i, width in enumerate(col_widths, 1):
        ws4.column_dimensions[get_column_letter(i)].width = width

    # Sheet 5: Risk Register
    ws5 = wb.create_sheet("Risk Register")

    risk_headers = ["ID", "Risk Description", "Category", "Impact", "Likelihood",
                   "Risk Score", "Mitigation Strategy", "Owner", "Status", "Notes"]

    for col, header in enumerate(risk_headers, 1):
        cell = ws5.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill

    risks = [
        ["R001", "Incomplete or inaccurate data from source systems", "Data", "High", "Medium", "6", "Early gap identification, multiple source triangulation", "BA", "Open", ""],
        ["R002", "Key stakeholders unavailable for interviews", "Resource", "Medium", "High", "6", "Flexible scheduling, backup interviewees, async options", "PM", "Open", ""],
        ["R003", "Scope creep from additional application discovery", "Scope", "High", "Medium", "6", "Clear scope definition, change control process", "Lead", "Open", ""],
        ["R004", "Executive engagement and decision delays", "Governance", "High", "Low", "3", "Regular updates, early preview, decision escalation", "Lead", "Open", ""],
        ["R005", "Tool access or technical issues", "Technical", "Medium", "Low", "2", "Early setup, backup procedures, IT support", "PM", "Open", ""],
        ["R006", "Resistance to recommendations", "Change", "High", "Medium", "6", "Early stakeholder involvement, validation sessions", "Lead", "Open", ""],
    ]

    for row_idx, risk in enumerate(risks, 2):
        for col_idx, value in enumerate(risk, 1):
            cell = ws5.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border

    col_widths = [6, 45, 12, 10, 10, 10, 45, 10, 10, 30]
    for i, width in enumerate(col_widths, 1):
        ws5.column_dimensions[get_column_letter(i)].width = width

    # Save workbook
    filepath = os.path.join(output_dir, "PROJECT_PLAN_DETAILED.xlsx")
    wb.save(filepath)
    print(f"Created: {filepath}")


def create_application_data_template():
    """Create APPLICATION_DATA_INPUT_TEMPLATE.xlsx."""
    wb = Workbook()

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="002B5B", end_color="002B5B", fill_type="solid")
    instruction_fill = PatternFill(start_color="E8F4FD", end_color="E8F4FD", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Sheet 1: Instructions
    ws = wb.active
    ws.title = "Instructions"

    instructions = [
        ["APPLICATION DATA INPUT TEMPLATE", ""],
        ["", ""],
        ["Purpose", "This template is used to collect application data for portfolio rationalization analysis."],
        ["", ""],
        ["Instructions:", ""],
        ["1.", "Fill in the 'Data Entry' sheet with information for each application"],
        ["2.", "Refer to 'Field Descriptions' for guidance on each field"],
        ["3.", "Use 'Scoring Reference' to understand how data maps to scores"],
        ["4.", "Leave fields blank if data is not available"],
        ["5.", "Use consistent formatting (dates as YYYY-MM-DD)"],
        ["", ""],
        ["Data Sources:", ""],
        ["", "- CMDB or Asset Management System"],
        ["", "- Financial/Accounting Systems"],
        ["", "- APM/Monitoring Tools"],
        ["", "- Service Desk/ITSM"],
        ["", "- HR Systems (for FTE data)"],
        ["", ""],
        ["Contact:", "For questions, contact your project team"],
    ]

    for row_idx, row in enumerate(instructions, 1):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                cell.font = Font(bold=True, size=16)
            elif row_idx in [3, 5, 12]:
                cell.font = Font(bold=True)

    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 80

    # Sheet 2: Data Entry
    ws2 = wb.create_sheet("Data Entry")

    data_headers = [
        "Application ID", "Application Name", "Description", "Business Owner",
        "Technical Owner", "Department", "Business Criticality", "User Count",
        "Technology Stack", "Hosting Model", "Vendor", "Version",
        "Install Date", "Last Major Update", "End of Support Date",
        "Annual License Cost", "Annual Infrastructure Cost", "Annual Support Cost",
        "FTE Support", "Monthly Incidents", "Monthly Availability %",
        "Integration Count", "Data Classification", "Compliance Requirements",
        "Replacement Planned", "Notes"
    ]

    for col, header in enumerate(data_headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    # Add sample rows
    sample_data = [
        ["APP001", "Enterprise ERP", "Core financial and operations system", "John Smith",
         "Sarah Johnson", "Finance", "Critical", "500", "SAP", "On-Premise", "SAP",
         "S/4HANA 2021", "2018-01-15", "2023-06-01", "2028-12-31", "250000", "150000",
         "100000", "3.5", "15", "99.5", "25", "Confidential", "SOX, GDPR", "No", ""],
        ["APP002", "CRM System", "Customer relationship management", "Mike Chen",
         "Lisa Wang", "Sales", "High", "200", "Salesforce", "Cloud/SaaS", "Salesforce",
         "Enterprise", "2020-03-01", "2024-01-15", "N/A", "180000", "0", "20000",
         "1.0", "5", "99.9", "12", "Confidential", "GDPR", "No", ""],
    ]

    for row_idx, row in enumerate(sample_data, 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border

    # Add data validations
    criticality_validation = DataValidation(
        type="list",
        formula1='"Critical,High,Medium,Low"',
        allow_blank=True
    )
    ws2.add_data_validation(criticality_validation)
    criticality_validation.add('G2:G1000')

    hosting_validation = DataValidation(
        type="list",
        formula1='"On-Premise,Cloud/SaaS,Cloud/IaaS,Cloud/PaaS,Hybrid,Hosted"',
        allow_blank=True
    )
    ws2.add_data_validation(hosting_validation)
    hosting_validation.add('J2:J1000')

    data_class_validation = DataValidation(
        type="list",
        formula1='"Public,Internal,Confidential,Restricted"',
        allow_blank=True
    )
    ws2.add_data_validation(data_class_validation)
    data_class_validation.add('W2:W1000')

    # Set column widths
    for col in range(1, len(data_headers) + 1):
        ws2.column_dimensions[get_column_letter(col)].width = 15

    # Sheet 3: Field Descriptions
    ws3 = wb.create_sheet("Field Descriptions")

    field_descriptions = [
        ["Field Name", "Description", "Data Type", "Required", "Example"],
        ["Application ID", "Unique identifier for the application", "Text", "Yes", "APP001"],
        ["Application Name", "Common name of the application", "Text", "Yes", "Enterprise ERP"],
        ["Description", "Brief description of application purpose", "Text", "Yes", "Core financial system"],
        ["Business Owner", "Primary business stakeholder", "Text", "Yes", "John Smith"],
        ["Technical Owner", "Primary technical contact", "Text", "Yes", "Sarah Johnson"],
        ["Department", "Primary department using the app", "Text", "Yes", "Finance"],
        ["Business Criticality", "Impact level if unavailable", "List", "Yes", "Critical/High/Medium/Low"],
        ["User Count", "Number of active users", "Number", "Yes", "500"],
        ["Technology Stack", "Primary technology/platform", "Text", "Yes", "SAP, .NET, Java"],
        ["Hosting Model", "Where application is hosted", "List", "Yes", "On-Premise/Cloud"],
        ["Vendor", "Software vendor or Internal", "Text", "Yes", "SAP"],
        ["Version", "Current version number", "Text", "No", "S/4HANA 2021"],
        ["Install Date", "Original installation date", "Date", "No", "2018-01-15"],
        ["Last Major Update", "Date of last major upgrade", "Date", "No", "2023-06-01"],
        ["End of Support Date", "Vendor support end date", "Date", "No", "2028-12-31"],
        ["Annual License Cost", "Yearly licensing fees", "Currency", "Yes", "250000"],
        ["Annual Infrastructure Cost", "Hardware, hosting, cloud costs", "Currency", "Yes", "150000"],
        ["Annual Support Cost", "Maintenance and support fees", "Currency", "Yes", "100000"],
        ["FTE Support", "Full-time equivalents for support", "Number", "Yes", "3.5"],
        ["Monthly Incidents", "Average incidents per month", "Number", "No", "15"],
        ["Monthly Availability %", "Average uptime percentage", "Percentage", "No", "99.5"],
        ["Integration Count", "Number of connected systems", "Number", "No", "25"],
        ["Data Classification", "Data sensitivity level", "List", "No", "Confidential"],
        ["Compliance Requirements", "Regulatory requirements", "Text", "No", "SOX, GDPR"],
        ["Replacement Planned", "Is replacement already planned", "Yes/No", "No", "No"],
        ["Notes", "Additional information", "Text", "No", ""],
    ]

    for row_idx, row in enumerate(field_descriptions, 1):
        for col_idx, value in enumerate(row, 1):
            cell = ws3.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                cell.font = header_font
                cell.fill = header_fill
            else:
                cell.border = thin_border

    col_widths = [25, 45, 12, 10, 25]
    for i, width in enumerate(col_widths, 1):
        ws3.column_dimensions[get_column_letter(i)].width = width

    # Sheet 4: Scoring Reference
    ws4 = wb.create_sheet("Scoring Reference")

    scoring_data = [
        ["SCORING REFERENCE", "", "", ""],
        ["", "", "", ""],
        ["Business Criticality Mapping", "", "", ""],
        ["Value", "Score", "Description", ""],
        ["Critical", "10", "Mission-critical, immediate business impact", ""],
        ["High", "7", "Important, significant impact within hours", ""],
        ["Medium", "4", "Moderate impact, workarounds available", ""],
        ["Low", "1", "Minimal impact, convenience application", ""],
        ["", "", "", ""],
        ["Technology Age Scoring", "", "", ""],
        ["Years Since Major Update", "Score", "Rating", ""],
        ["< 1 year", "10", "Current", ""],
        ["1-2 years", "8", "Recent", ""],
        ["2-3 years", "6", "Aging", ""],
        ["3-5 years", "4", "Old", ""],
        ["> 5 years", "2", "Legacy", ""],
        ["", "", "", ""],
        ["Availability Scoring", "", "", ""],
        ["Availability %", "Score", "Rating", ""],
        [">= 99.9%", "10", "Excellent", ""],
        ["99.5-99.9%", "8", "Good", ""],
        ["99.0-99.5%", "6", "Acceptable", ""],
        ["98.0-99.0%", "4", "Poor", ""],
        ["< 98%", "2", "Critical", ""],
    ]

    for row_idx, row in enumerate(scoring_data, 1):
        for col_idx, value in enumerate(row, 1):
            cell = ws4.cell(row=row_idx, column=col_idx, value=value)
            if row_idx in [1]:
                cell.font = Font(bold=True, size=14)
            elif row_idx in [3, 10, 18]:
                cell.font = Font(bold=True)
                cell.fill = instruction_fill

    ws4.column_dimensions['A'].width = 25
    ws4.column_dimensions['B'].width = 10
    ws4.column_dimensions['C'].width = 40

    # Save workbook
    filepath = os.path.join(output_dir, "APPLICATION_DATA_INPUT_TEMPLATE.xlsx")
    wb.save(filepath)
    print(f"Created: {filepath}")


if __name__ == "__main__":
    create_project_plan_detailed()
    create_application_data_template()
    print("All Excel templates created successfully!")
